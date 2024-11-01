from datetime import datetime, timedelta

from aiogram import types
from aiogram.dispatcher import FSMContext

from keyboards.default.companies_keyboards import all_companies_default_keyboard
from keyboards.default.go_to_registration import go_registration_default_keyboard
from keyboards.default.menu_keyboards import back_to_menu
from loader import dp, db, bot

import os
import tempfile
from openpyxl.styles import Alignment
import openpyxl

from states.download_states import DownloadQualityState


async def download_thirty_day_qualities(company_id):
    qualities = await db.select_qualities_by_company_id(company_id=company_id)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = "Firma nomi"
    worksheet['C1'] = 'Zakaz Kolichestvo'
    worksheet['D1'] = 'Dostavleno Kolichestvo'
    worksheet['E1'] = 'Protsent akkuratnost'
    worksheet['F1'] = 'Vremya podog'
    worksheet['G1'] = 'Vremya fakt dostavka'
    worksheet['H1'] = 'Vremya akkuratnost'
    worksheet['I1'] = 'Kalichestvo tovara'
    worksheet['J1'] = 'Sredniy protsent akkuratnost'

    for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1']:
        worksheet[cell].alignment = Alignment(horizontal='center')

    worksheet.cell(row=1, column=1, value='T/r')
    worksheet.cell(row=1, column=2, value='Firma nomi')
    worksheet.cell(row=1, column=3, value='Zakaz Kolichestvo')
    worksheet.cell(row=1, column=4, value='Dostavleno Kolichestvo')
    worksheet.cell(row=1, column=5, value='Protsent akkuratnost')
    worksheet.cell(row=1, column=6, value='Vremya podog')
    worksheet.cell(row=1, column=7, value='Vremya fakt dostavka')
    worksheet.cell(row=1, column=8, value='Vremya akkuratnost')
    worksheet.cell(row=1, column=9, value='Kalichestvo tovara')
    worksheet.cell(row=1, column=10, value="Sredniy protsent akkuratnost")

    companies = await db.select_companies(id=company_id)
    company = companies[0]
    company_name = company['name']

    tr = 0
    for quality in qualities:
        tr += 1

        ordered_quantity = quality['ordered_quantity']
        delivered_quantity = quality['delivered_quantity']
        percent_products = quality['percent_products']
        agreed_time = quality['agreed_time'] + timedelta(hours=5)
        delivered_time = quality['delivered_time'] + timedelta(hours=5)
        description_time = quality['description_time']
        quality_description = quality['quality_description']
        average_percentage = quality['average_percentage']

        worksheet.cell(row=tr + 1, column=1, value=tr).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=2, value=company_name).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=3, value=ordered_quantity).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=4, value=delivered_quantity).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=5, value=f"{percent_products} %").alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=6, value=agreed_time.strftime("%d.%m.%Y  %H:%M")).alignment = Alignment(
            horizontal='center')
        worksheet.cell(row=tr + 1, column=7, value=delivered_time.strftime("%d.%m.%Y  %H:%M")).alignment = Alignment(
            horizontal='center')
        worksheet.cell(row=tr + 1, column=8, value=description_time).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=9, value=quality_description).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=10, value=f"{average_percentage} %").alignment = Alignment(
            horizontal='center')

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, 'Bizbop_akkuratnost_postavok.xlsx')
    workbook.save(file_path)

    return temp_dir


@dp.message_handler(text="📥 Excel fayl yuklab olish", state="*")
async def start_getting_company_id(message: types.Message, state: FSMContext):
    try:
        await state.finish()
    except:
        pass
    user_telegram_id = message.from_user.id
    users = await db.select_users(telegram_id=user_telegram_id)
    if not users:
        await message.answer(text="🚫 Sizda botdan foydalanish uchun ruxsat mavjud emas,\n"
                                  "📝 Botdan foydalanish uchun ro'yxatdan o'tishingiz kerak 👇",
                             reply_markup=go_registration_default_keyboard)
        return
    user = users[0]
    user_role = user['role']
    if user_role != "admin":
        await message.answer(text="🚫 Sizda botdan foydalanish uchun ruxsat mavjud emas", reply_markup=back_to_menu)
        return
    markup = await all_companies_default_keyboard()
    await message.answer(text="🏢 Hisobotni yuklab olish uchun Kompaniyani tanlang",
                         reply_markup=markup)
    await DownloadQualityState.company_id.set()


@dp.message_handler(state=DownloadQualityState.company_id)
async def download_qualities_function(message: types.Message, state: FSMContext):
    company_name = message.text
    companies = await db.select_companies(name=company_name)

    if not companies:
        await message.answer(text="⚠️ Bunday firma topilmadi", reply_markup=back_to_menu)
        return
    company_id = companies[0]['id']
    await state.update_data(company_id=company_id)

    temp_dir = await download_thirty_day_qualities(company_id=company_id)

    with open(os.path.join(temp_dir, 'Bizbop_akkuratnost_postavok.xlsx'), 'rb') as file:
        await message.answer_document(document=file)

    os.remove(os.path.join(temp_dir, 'Bizbop_akkuratnost_postavok.xlsx'))
